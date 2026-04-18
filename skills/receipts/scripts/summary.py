#!/usr/bin/env python3
"""Show expense summary from the receipts Excel file."""

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed", file=sys.stderr)
    sys.exit(1)

RCLONE_CONFIG = "/data/.config/rclone/rclone.conf"
LOCAL_DIR = "/data/.openclaw/workspace/receipts"
YEAR = datetime.now().strftime("%Y")
EXCEL_NAME = f"Expenses_{YEAR}.xlsx"
LOCAL_EXCEL = os.path.join(LOCAL_DIR, EXCEL_NAME)
DRIVE_ACCOUNTANT = "gdrive:Accountant/"


def main():
    parser = argparse.ArgumentParser(description="Expense summary")
    parser.add_argument("--month", help="Filter by month (YYYY-MM)")
    parser.add_argument("--year", help="Full year summary (YYYY)")
    parser.add_argument("--category", help="Filter by category")
    args = parser.parse_args()

    os.makedirs(LOCAL_DIR, exist_ok=True)

    # Download latest Excel from Drive
    subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy",
                     DRIVE_ACCOUNTANT + EXCEL_NAME, LOCAL_DIR],
                    capture_output=True)

    if not os.path.exists(LOCAL_EXCEL):
        print("No expense file found. Process some receipts first.")
        return

    wb = openpyxl.load_workbook(LOCAL_EXCEL)
    ws = wb["Transactions"]

    results = []
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value or ""
        vendor = ws.cell(row=row, column=2).value or ""
        desc = ws.cell(row=row, column=3).value or ""
        category = ws.cell(row=row, column=4).value or ""
        subtotal = ws.cell(row=row, column=5).value or 0
        gst = ws.cell(row=row, column=6).value or 0
        pst = ws.cell(row=row, column=7).value or 0
        total = ws.cell(row=row, column=8).value or 0

        # Apply filters
        if args.month and not str(date_val).startswith(args.month):
            continue
        if args.year and not str(date_val).startswith(args.year):
            continue
        if args.category and args.category.lower() not in category.lower():
            continue

        results.append({
            "date": str(date_val),
            "vendor": vendor,
            "description": desc,
            "category": category,
            "subtotal": subtotal,
            "gst_hst": gst,
            "pst": pst,
            "total": total,
        })

    if not results:
        print("No transactions found matching your filters.")
        return

    # Summary
    total_spent = sum(r["total"] for r in results)
    total_gst = sum(r["gst_hst"] for r in results)
    total_pst = sum(r["pst"] for r in results)

    # By category
    by_cat = {}
    for r in results:
        cat = r["category"]
        if cat not in by_cat:
            by_cat[cat] = {"total": 0, "count": 0}
        by_cat[cat]["total"] += r["total"]
        by_cat[cat]["count"] += 1

    filter_desc = ""
    if args.month:
        filter_desc = f" for {args.month}"
    elif args.year:
        filter_desc = f" for {args.year}"
    if args.category:
        filter_desc += f" in {args.category}"

    print(f"=== Expense Summary{filter_desc} ===\n")
    print(f"Transactions: {len(results)}")
    print(f"Total Spent:  ${total_spent:,.2f}")
    print(f"GST/HST Paid: ${total_gst:,.2f}")
    print(f"PST Paid:     ${total_pst:,.2f}")
    print(f"\n--- By Category ---")
    for cat in sorted(by_cat.keys()):
        data = by_cat[cat]
        print(f"  {cat}: ${data['total']:,.2f} ({data['count']} transactions)")

    print(f"\n--- Recent Transactions ---")
    for r in results[-10:]:
        print(f"  {r['date']} | {r['vendor'][:25]:<25} | ${r['total']:>8,.2f} | {r['category']}")


if __name__ == "__main__":
    main()
