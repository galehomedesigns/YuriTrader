#!/usr/bin/env python3
"""Process a single receipt image with quality checks and feedback."""

import base64
import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path

RCLONE_CONFIG = (os.path.expanduser("~/.config/rclone/rclone.conf") if not os.path.exists("/data/.openclaw") else "/data/.config/rclone/rclone.conf")
DRIVE_PROCESSED = "gdrive:Receipts/Processed/"
DRIVE_ACCOUNTANT = "gdrive:Accountant/"
LOCAL_DIR = (os.path.join(os.environ.get("OPENCLAW_ROOT", "/home/tonygale/openclaw"), "workspace", "receipts") if not os.path.exists("/data/.openclaw") else "/data/.openclaw/workspace/receipts")
YEAR = datetime.now().strftime("%Y")
EXCEL_NAME = f"Expenses_{YEAR}.xlsx"
LOCAL_EXCEL = os.path.join(LOCAL_DIR, EXCEL_NAME)
PROCESSED_LOG = os.path.join(LOCAL_DIR, "processed.json")

CRA_CATEGORIES = [
    "Advertising & Marketing", "Meals & Entertainment", "Office Supplies",
    "Professional Fees", "Rent & Lease", "Telephone & Internet", "Travel",
    "Vehicle Expenses", "Software & Subscriptions", "Equipment & Assets",
    "Insurance", "Training & Education", "Bank & Interest Charges",
    "Shipping & Delivery", "Subcontractors", "Other",
]


def analyze_receipt(image_path):
    """Send receipt image to local qwen2.5vl:72b (via Ollama) for vision OCR
    + structured extraction. Returns parsed JSON dict."""
    with open(image_path, "rb") as f:
        image_data = base64.b64encode(f.read()).decode()

    categories_str = ", ".join(CRA_CATEGORIES)
    prompt = f"""Analyze this image. First determine what it is, then assess quality.

Return JSON in this exact format:
{{
  "is_receipt": true/false,
  "image_quality": "good" | "fair" | "poor" | "unreadable",
  "quality_issues": ["list of specific issues, e.g. 'total amount is blurry', 'date is cut off', 'glare on bottom half'"],
  "confidence": 0.0 to 1.0,
  "vendor": "Store/business name or null if unreadable",
  "date": "YYYY-MM-DD or null if unreadable",
  "dates_found": ["YYYY-MM-DD", "..."],
  "subtotal": 0.00,
  "tax_gst_hst": 0.00,
  "tax_pst": 0.00,
  "total": 0.00,
  "currency": "CAD",
  "payment_method": "cash/credit/debit/other or null",
  "description": "Brief description of purchase",
  "category": "One of: {categories_str}",
  "items": ["item1", "item2"],
  "what_is_it": "Description if not a receipt (e.g. 'a photo of a dog', 'a business card')"
}}

Rules:
- If the image is NOT a receipt, set is_receipt to false and describe what it is in what_is_it
- If the image IS a receipt but parts are unreadable, set those fields to null and list specific issues in quality_issues
- confidence: how confident you are in the extracted data (0.0 = can't read anything, 1.0 = crystal clear)
- image_quality: "good" (all text clear), "fair" (mostly readable, some issues), "poor" (significant parts unreadable), "unreadable" (can't extract meaningful data)
- Use CAD unless another currency is clearly shown
- For Canadian receipts, separate GST/HST and PST if shown
- dates_found: list EVERY distinct transaction/print date you can see on the receipt (header, footer, timestamp, payment line, etc.) in YYYY-MM-DD form. Exclude "best before"/expiry/coupon dates. If a date appears multiple times in the same form, list it once. Most receipts show the same date 2-3 times — use this to cross-check yourself. If they disagree, set "date" to null and add "conflicting dates: X vs Y" to quality_issues.
- Beware misread year digits. "2020" vs "2026" or "2025" vs "2026" are common OCR failures on the 0/5/6 character. If two dates differ only in year, trust the one that appears more often or is most legible.
- Return ONLY valid JSON, no markdown fences, no other text."""

    ollama_url = os.environ.get("OLLAMA_BASE_URL", "http://127.0.0.1:11434")
    payload = {
        "model": "qwen2.5vl:72b",
        "prompt": prompt,
        "images": [image_data],
        "stream": False,
        "format": "json",
        "keep_alive": "30m",
        "options": {"temperature": 0.1, "num_ctx": 8192, "num_predict": 2048},
    }

    import urllib.request
    headers = {"Content-Type": "application/json"}
    api_key = os.environ.get("OLLAMA_API_KEY")
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    url = f"{ollama_url}/api/generate"
    req = urllib.request.Request(url, data=json.dumps(payload).encode(),
                                 headers=headers)

    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            result = json.loads(resp.read())
        text = (result.get("response") or "").strip()
        # Strip markdown fences if the model emitted any despite format=json
        if text.startswith("```"):
            text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3]
        return json.loads(text.strip())
    except Exception as e:
        return {"error": f"Ollama vision error: {e}"}


def main():
    if len(sys.argv) < 2:
        print("Usage: process_single.py <image_path> [--save]", file=sys.stderr)
        sys.exit(1)

    image_path = sys.argv[1]
    should_save = "--save" in sys.argv

    if not os.path.exists(image_path):
        print(json.dumps({"error": f"File not found: {image_path}"}))
        sys.exit(1)

    os.makedirs(LOCAL_DIR, exist_ok=True)

    result = analyze_receipt(image_path)

    if "error" in result:
        print(json.dumps(result, indent=2))
        sys.exit(1)

    # Cross-check: if multiple dates appeared on the receipt and they disagree,
    # null the date and flag it. Receipts almost always print the same date 2-3
    # times; a mismatch points to an OCR error.
    dates_found = result.get("dates_found") or []
    distinct = sorted({d for d in dates_found if d})
    if len(distinct) > 1:
        result["date"] = None
        issues = result.setdefault("quality_issues", [])
        issues.append(f"conflicting dates on receipt: {', '.join(distinct)}")
        if result.get("image_quality") in ("good", "fair"):
            result["image_quality"] = "poor"

    # Not a receipt
    if not result.get("is_receipt", False):
        output = {
            "status": "not_a_receipt",
            "message": f"This doesn't appear to be a receipt. It looks like: {result.get('what_is_it', 'unknown')}",
        }
        print(json.dumps(output, indent=2))
        sys.exit(0)

    # Quality check
    quality = result.get("image_quality", "unknown")
    confidence = result.get("confidence", 0)
    issues = result.get("quality_issues", [])

    if quality == "unreadable" or confidence < 0.3:
        output = {
            "status": "unreadable",
            "message": "This receipt image is too blurry or damaged to read.",
            "issues": issues,
            "suggestion": "Please retake the photo. Tips: hold steady, ensure good lighting, capture the entire receipt including the total at the bottom.",
        }
        print(json.dumps(output, indent=2))
        sys.exit(0)

    if quality == "poor" or confidence < 0.6:
        output = {
            "status": "poor_quality",
            "message": "I can partially read this receipt but some details are unclear.",
            "issues": issues,
            "partial_data": {
                "vendor": result.get("vendor"),
                "date": result.get("date"),
                "total": result.get("total"),
                "category": result.get("category"),
            },
            "suggestion": "I extracted what I could (above). You can confirm these details and I'll save them, or retake the photo for better accuracy.",
        }
        print(json.dumps(output, indent=2))
        sys.exit(0)

    # Good or fair quality — show extracted data
    output = {
        "status": "success",
        "quality": quality,
        "confidence": confidence,
        "data": {
            "vendor": result.get("vendor"),
            "date": result.get("date"),
            "subtotal": result.get("subtotal"),
            "tax_gst_hst": result.get("tax_gst_hst"),
            "tax_pst": result.get("tax_pst"),
            "total": result.get("total"),
            "currency": result.get("currency", "CAD"),
            "payment_method": result.get("payment_method"),
            "description": result.get("description"),
            "category": result.get("category"),
            "items": result.get("items", []),
        },
    }

    if issues:
        output["quality_notes"] = issues

    # Check for duplicates
    try:
        sys.path.insert(0, "/home/tonygale/openclaw/skills/receipts/scripts")
        from dedup_check import find_duplicate
        dup = find_duplicate(output["data"].get("vendor"), output["data"].get("date"), output["data"].get("total"))
        if dup:
            output["duplicate_warning"] = f"Possible duplicate: {dup['vendor']} on {dup['date']} for ${dup['total']:.2f} (row {dup['row']}, category: {dup['category']}). Already in spreadsheet."
    except Exception:
        pass

    if not should_save:
        output["next_step"] = "Ask Tony to confirm, then re-run with --save to record it."
        print(json.dumps(output, indent=2))
        sys.exit(0)

    # Save to Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        output["save_error"] = "openpyxl not installed"
        print(json.dumps(output, indent=2))
        sys.exit(1)

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Download existing Excel from Drive if not local
    if not os.path.exists(LOCAL_EXCEL):
        subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy",
                         DRIVE_ACCOUNTANT + EXCEL_NAME, LOCAL_DIR], capture_output=True)

    if os.path.exists(LOCAL_EXCEL):
        wb = openpyxl.load_workbook(LOCAL_EXCEL)
    else:
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        headers = ["Date", "Vendor", "Description", "Category", "Subtotal", "GST/HST",
                    "PST", "Total", "Currency", "Payment Method", "Receipt File"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        ws2 = wb.create_sheet("Category Summary")
        headers2 = ["Category", "Total Spent", "GST/HST Paid", "PST Paid", "Transaction Count"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for row, cat in enumerate(CRA_CATEGORIES, 2):
            ws2.cell(row=row, column=1, value=cat)

    ws = wb["Transactions"]
    data = output["data"]
    row = ws.max_row + 1
    filename = os.path.basename(image_path)

    ws.cell(row=row, column=1, value=data.get("date", "")).border = THIN_BORDER
    ws.cell(row=row, column=2, value=data.get("vendor", "")).border = THIN_BORDER
    ws.cell(row=row, column=3, value=data.get("description", "")).border = THIN_BORDER
    ws.cell(row=row, column=4, value=data.get("category", "Other")).border = THIN_BORDER
    ws.cell(row=row, column=5, value=data.get("subtotal", 0)).border = THIN_BORDER
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=data.get("tax_gst_hst", 0)).border = THIN_BORDER
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=7, value=data.get("tax_pst", 0)).border = THIN_BORDER
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    ws.cell(row=row, column=8, value=data.get("total", 0)).border = THIN_BORDER
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    ws.cell(row=row, column=9, value=data.get("currency", "CAD")).border = THIN_BORDER
    ws.cell(row=row, column=10, value=data.get("payment_method", "")).border = THIN_BORDER
    # Column 11 is "Card Ref" — left empty (Gemini doesn't extract card numbers).
    # Filename goes in column 12 "Receipt/Gmail Ref".
    ws.cell(row=row, column=12, value=filename).border = THIN_BORDER

    wb.save(LOCAL_EXCEL)

    # Upload Excel to Drive
    subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy", LOCAL_EXCEL, DRIVE_ACCOUNTANT],
                    capture_output=True)

    # Backup receipt image to Drive
    subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy", image_path, DRIVE_PROCESSED[:-1]],
                    capture_output=True)

    output["saved"] = True
    output["message"] = f"Receipt saved to {EXCEL_NAME} and backed up to Google Drive."
    print(json.dumps(output, indent=2))


if __name__ == "__main__":
    main()
