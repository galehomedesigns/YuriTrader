#!/usr/bin/env python3
"""Process receipt images from Google Drive using Gemini vision API."""

import argparse
import base64
import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RCLONE_CONFIG = (os.path.expanduser("~/.config/rclone/rclone.conf") if not os.path.exists("/data/.openclaw") else "/data/.config/rclone/rclone.conf")
DRIVE_RECEIPTS = "gdrive:Receipts/"
DRIVE_PROCESSED = "gdrive:Receipts/Processed/"
DRIVE_ACCOUNTANT = "gdrive:Accountant/"
LOCAL_DIR = (os.path.join(os.environ.get("OPENCLAW_ROOT", "/home/tonygale/openclaw"), "workspace", "receipts") if not os.path.exists("/data/.openclaw") else "/data/.openclaw/workspace/receipts")
YEAR = datetime.now().strftime("%Y")
EXCEL_NAME = f"Expenses_{YEAR}.xlsx"
LOCAL_EXCEL = os.path.join(LOCAL_DIR, EXCEL_NAME)
PROCESSED_LOG = os.path.join(LOCAL_DIR, "processed.json")

CRA_CATEGORIES = [
    "Advertising & Marketing",
    "Meals & Entertainment",
    "Office Supplies",
    "Professional Fees",
    "Rent & Lease",
    "Telephone & Internet",
    "Travel",
    "Vehicle Expenses",
    "Software & Subscriptions",
    "Equipment & Assets",
    "Insurance",
    "Training & Education",
    "Bank & Interest Charges",
    "Shipping & Delivery",
    "Subcontractors",
    "Other",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def rclone(*args):
    cmd = ["rclone", "--config", RCLONE_CONFIG] + list(args)
    result = subprocess.run(cmd, capture_output=True, text=True)
    return result.stdout.strip(), result.stderr.strip(), result.returncode


def list_new_receipts():
    """List image files in the Receipts folder (non-recursively)."""
    out, err, rc = rclone("lsf", "--files-only", DRIVE_RECEIPTS)
    if rc != 0:
        return []
    files = []
    for filename in out.split("\n"):
        filename = filename.strip()
        if not filename:
            continue
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext in ("jpg", "jpeg", "png", "heic", "webp", "pdf"):
            files.append(filename)
    return files


def load_processed():
    """Load list of already processed filenames."""
    if os.path.exists(PROCESSED_LOG):
        with open(PROCESSED_LOG) as f:
            return json.load(f)
    return []


def save_processed(processed):
    with open(PROCESSED_LOG, "w") as f:
        json.dump(processed, f, indent=2)


def download_receipt(filename):
    """Download a receipt from Drive to local dir."""
    local_path = os.path.join(LOCAL_DIR, "incoming", filename)
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    _, err, rc = rclone("copy", DRIVE_RECEIPTS + filename, os.path.dirname(local_path))
    if rc != 0:
        print(f"Error downloading {filename}: {err}", file=sys.stderr)
        return None
    return local_path


def analyze_receipt_with_gemini(image_path):
    """Send receipt image to local qwen2.5vl:72b (via Ollama) for vision OCR
    + structured extraction. (Function name kept for caller compatibility.)"""
    with open(image_path, "rb") as f:
        image_data = base64.b64encode(f.read()).decode()

    categories_str = ", ".join(CRA_CATEGORIES)
    prompt = f"""Analyze this receipt image and extract the following information as JSON:
{{
  "vendor": "Store/business name",
  "date": "YYYY-MM-DD",
  "dates_found": ["YYYY-MM-DD", "..."],
  "subtotal": 0.00,
  "tax_gst_hst": 0.00,
  "tax_pst": 0.00,
  "total": 0.00,
  "currency": "CAD",
  "payment_method": "cash/credit/debit/other",
  "description": "Brief description of purchase",
  "category": "One of: {categories_str}",
  "items": ["item1", "item2"]
}}

Rules:
- Use CAD unless another currency is clearly shown
- For Canadian receipts, separate GST/HST and PST if shown
- Choose the most appropriate CRA category from the list
- If date is unclear, use null
- dates_found: list EVERY distinct transaction/print date visible on the receipt (header, footer, timestamp, payment line, etc.) in YYYY-MM-DD form. Exclude "best before"/expiry/coupon dates. Most receipts print the same date 2-3 times — use this to cross-check yourself. If they disagree, set "date" to null.
- Beware misread year digits. "2020" vs "2026" or "2025" vs "2026" are common OCR failures on 0/5/6. If two dates differ only in year, trust the one that appears more often or is most legible.
- Return ONLY valid JSON, no markdown fences, no other text."""

    ollama_url = os.environ.get("OLLAMA_BASE_URL", "http://127.0.0.1:11434")
    payload = {
        "model": "qwen2.5vl:72b",
        "prompt": prompt,
        "images": [image_data],
        "stream": False,
        "format": "json",
        "keep_alive": "30m",
        "options": {"temperature": 0.1, "num_ctx": 8192, "num_predict": 2048},
    }

    import urllib.request
    headers = {"Content-Type": "application/json"}
    api_key = os.environ.get("OLLAMA_API_KEY")
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    url = f"{ollama_url}/api/generate"
    req = urllib.request.Request(url, data=json.dumps(payload).encode(),
                                 headers=headers)

    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            result = json.loads(resp.read())
        text = (result.get("response") or "").strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
        if not text.endswith("}"):
            print("Warning: Ollama response may be truncated", file=sys.stderr)
            return None
        parsed = json.loads(text)
        dates_found = parsed.get("dates_found") or []
        distinct = sorted({d for d in dates_found if d})
        if len(distinct) > 1:
            print(f"  date conflict: {distinct} — nulling date", file=sys.stderr)
            parsed["date"] = None
        return parsed
    except Exception as e:
        print(f"Ollama vision error: {e}", file=sys.stderr)
        return None


def move_to_processed(filename):
    """Move receipt to Processed subfolder on Drive."""
    rclone("mkdir", DRIVE_PROCESSED)
    rclone("moveto", DRIVE_RECEIPTS + filename, DRIVE_PROCESSED + filename)


def init_excel():
    """Create or load the Excel spreadsheet."""
    if os.path.exists(LOCAL_EXCEL):
        return openpyxl.load_workbook(LOCAL_EXCEL)

    wb = openpyxl.Workbook()

    # Transactions sheet
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Vendor", "Description", "Category", "Subtotal", "GST/HST",
               "PST", "Total", "Currency", "Payment Method", "Receipt File"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 30

    # Summary sheet
    ws2 = wb.create_sheet("Category Summary")
    headers2 = ["Category", "Total Spent", "GST/HST Paid", "PST Paid", "Transaction Count"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    for row, cat in enumerate(CRA_CATEGORIES, 2):
        ws2.cell(row=row, column=1, value=cat)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 18

    wb.save(LOCAL_EXCEL)
    return wb


def add_transaction(wb, data, filename):
    """Add a receipt transaction to the spreadsheet."""
    ws = wb["Transactions"]
    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=data.get("date", "")).border = THIN_BORDER
    ws.cell(row=row, column=2, value=data.get("vendor", "")).border = THIN_BORDER
    ws.cell(row=row, column=3, value=data.get("description", "")).border = THIN_BORDER
    ws.cell(row=row, column=4, value=data.get("category", "Other")).border = THIN_BORDER
    ws.cell(row=row, column=5, value=data.get("subtotal", 0)).border = THIN_BORDER
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=data.get("tax_gst_hst", 0)).border = THIN_BORDER
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=7, value=data.get("tax_pst", 0)).border = THIN_BORDER
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    ws.cell(row=row, column=8, value=data.get("total", 0)).border = THIN_BORDER
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    ws.cell(row=row, column=9, value=data.get("currency", "CAD")).border = THIN_BORDER
    ws.cell(row=row, column=10, value=data.get("payment_method", "")).border = THIN_BORDER
    # Filename goes in column 12 "Receipt/Gmail Ref" (column 11 "Card Ref" is left empty).
    ws.cell(row=row, column=12, value=filename).border = THIN_BORDER

    wb.save(LOCAL_EXCEL)


def update_summary(wb):
    """Update the category summary sheet with totals."""
    ws_trans = wb["Transactions"]
    ws_sum = wb["Category Summary"]

    # Build totals by category
    totals = {}
    for row in range(2, ws_trans.max_row + 1):
        cat = ws_trans.cell(row=row, column=4).value or "Other"
        total = ws_trans.cell(row=row, column=8).value or 0
        gst = ws_trans.cell(row=row, column=6).value or 0
        pst = ws_trans.cell(row=row, column=7).value or 0
        if cat not in totals:
            totals[cat] = {"total": 0, "gst": 0, "pst": 0, "count": 0}
        totals[cat]["total"] += total
        totals[cat]["gst"] += gst
        totals[cat]["pst"] += pst
        totals[cat]["count"] += 1

    for row in range(2, len(CRA_CATEGORIES) + 2):
        cat = ws_sum.cell(row=row, column=1).value
        data = totals.get(cat, {"total": 0, "gst": 0, "pst": 0, "count": 0})
        ws_sum.cell(row=row, column=2, value=data["total"]).number_format = '#,##0.00'
        ws_sum.cell(row=row, column=3, value=data["gst"]).number_format = '#,##0.00'
        ws_sum.cell(row=row, column=4, value=data["pst"]).number_format = '#,##0.00'
        ws_sum.cell(row=row, column=5, value=data["count"])
        for col in range(2, 6):
            ws_sum.cell(row=row, column=col).border = THIN_BORDER

    # Grand total row
    grand_row = len(CRA_CATEGORIES) + 3
    ws_sum.cell(row=grand_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    for col, key in [(2, "total"), (3, "gst"), (4, "pst"), (5, "count")]:
        val = sum(t[key] for t in totals.values())
        ws_sum.cell(row=grand_row, column=col, value=val).font = Font(bold=True)
        if col < 5:
            ws_sum.cell(row=grand_row, column=col).number_format = '#,##0.00'

    wb.save(LOCAL_EXCEL)


def upload_excel():
    """Upload the Excel file to Google Drive Accountant folder."""
    rclone("copy", LOCAL_EXCEL, DRIVE_ACCOUNTANT)


def main():
    parser = argparse.ArgumentParser(description="Process receipts from Google Drive")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be processed")
    parser.add_argument("--reprocess", action="store_true", help="Reprocess all receipts")
    args = parser.parse_args()

    os.makedirs(LOCAL_DIR, exist_ok=True)
    os.makedirs(os.path.join(LOCAL_DIR, "incoming"), exist_ok=True)

    # Download existing Excel from Drive if we don't have it locally
    if not os.path.exists(LOCAL_EXCEL):
        rclone("copy", DRIVE_ACCOUNTANT + EXCEL_NAME, LOCAL_DIR)

    # List receipts
    all_files = list_new_receipts()
    processed = load_processed() if not args.reprocess else []
    new_files = [f for f in all_files if f not in processed]

    if not new_files:
        print("No new receipts to process.")
        return

    print(f"Found {len(new_files)} new receipt(s):")
    for f in new_files:
        print(f"  - {f}")

    if args.dry_run:
        print("\nDry run — no changes made.")
        return

    wb = init_excel()
    success_count = 0

    for filename in new_files:
        print(f"\nProcessing: {filename}")

        # Download
        local_path = download_receipt(filename)
        if not local_path or not os.path.exists(local_path):
            print(f"  Skipped: could not download")
            continue

        # Analyze with Gemini
        print(f"  Analyzing with Gemini...")
        data = analyze_receipt_with_gemini(local_path)
        if not data:
            print(f"  Skipped: Gemini analysis failed")
            continue

        print(f"  Vendor: {data.get('vendor', 'Unknown')}")
        print(f"  Date: {data.get('date', 'Unknown')}")
        print(f"  Total: ${data.get('total', 0):.2f} {data.get('currency', 'CAD')}")
        print(f"  Category: {data.get('category', 'Other')}")

        # Add to Excel
        add_transaction(wb, data, filename)
        processed.append(filename)

        # Move receipt to Processed folder
        move_to_processed(filename)

        # Cleanup local file
        if os.path.exists(local_path):
            os.remove(local_path)

        success_count += 1

    # Update summary and upload
    update_summary(wb)
    upload_excel()
    save_processed(processed)

    print(f"\nDone! Processed {success_count}/{len(new_files)} receipts.")
    print(f"Excel uploaded to Drive: Accountant/{EXCEL_NAME}")


if __name__ == "__main__":
    main()
