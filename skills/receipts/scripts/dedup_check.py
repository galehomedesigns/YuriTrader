#!/usr/bin/env python3
"""Check for duplicate expenses in the spreadsheet.

Shared module used by process_single.py, add_expense.py, and scan_email_receipts.py
to prevent duplicate entries based on vendor + date + total amount.
"""

import os
import subprocess

RCLONE_CONFIG = (os.path.expanduser("~/.config/rclone/rclone.conf") if not os.path.exists("/data/.openclaw") else "/data/.config/rclone/rclone.conf")
LOCAL_DIR = (os.path.join(os.environ.get("OPENCLAW_ROOT", "/home/tonygale/openclaw"), "workspace", "receipts") if not os.path.exists("/data/.openclaw") else "/data/.openclaw/workspace/receipts")
DRIVE_ACCOUNTANT = "gdrive:Accountant/"


def find_duplicate(vendor, date_str, total, excel_path=None):
    """Check if an expense with matching vendor+date+total already exists.

    Returns dict with match info if duplicate found, None otherwise.
    Matches are fuzzy on vendor name (case-insensitive, substring match).
    Matches are exact on total (within $0.01) and date.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    if excel_path is None:
        from datetime import datetime
        year = datetime.now().strftime("%Y")
        excel_path = os.path.join(LOCAL_DIR, f"Expenses_{year}.xlsx")

    # Download latest if not local
    if not os.path.exists(excel_path):
        excel_name = os.path.basename(excel_path)
        subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy",
                         DRIVE_ACCOUNTANT + excel_name, LOCAL_DIR], capture_output=True)

    if not os.path.exists(excel_path):
        return None

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Transactions"]
    except Exception:
        return None

    vendor_lower = (vendor or "").lower().strip()
    date_clean = str(date_str or "").strip()
    total_float = float(total or 0)

    for row in range(2, ws.max_row + 1):
        existing_date = str(ws.cell(row=row, column=1).value or "").strip()
        existing_vendor = str(ws.cell(row=row, column=2).value or "").lower().strip()
        existing_total = float(ws.cell(row=row, column=8).value or 0)

        # Date match (exact string)
        if existing_date != date_clean:
            continue

        # Total match (within $0.01)
        if abs(existing_total - total_float) > 0.01:
            continue

        # Vendor match (fuzzy — substring in either direction)
        if vendor_lower in existing_vendor or existing_vendor in vendor_lower:
            match = {
                "row": row,
                "date": existing_date,
                "vendor": ws.cell(row=row, column=2).value,
                "total": existing_total,
                "category": ws.cell(row=row, column=4).value,
            }
            wb.close()
            return match

    wb.close()
    return None
