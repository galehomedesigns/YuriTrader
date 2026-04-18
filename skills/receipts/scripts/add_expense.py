#!/usr/bin/env python3
"""Manually add an expense entry to the Expenses spreadsheet.

Usage:
    python3 add_expense.py --vendor "Uber" --total 25.50 --date "2026-03-07" \
        --category "Travel" --description "Uber ride downtown" \
        --gst 0 --pst 0 --payment "Visa" --source "Email receipt"
"""

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime

RCLONE_CONFIG = "/data/.config/rclone/rclone.conf"
LOCAL_DIR = "/data/.openclaw/workspace/receipts"
YEAR = datetime.now().strftime("%Y")
EXCEL_NAME = f"Expenses_{YEAR}.xlsx"
LOCAL_EXCEL = os.path.join(LOCAL_DIR, EXCEL_NAME)
DRIVE_ACCOUNTANT = "gdrive:Accountant/"

CRA_CATEGORIES = [
    "Advertising & Marketing", "Meals & Entertainment", "Office Supplies",
    "Professional Fees", "Rent & Lease", "Telephone & Internet", "Travel",
    "Vehicle Expenses", "Software & Subscriptions", "Equipment & Assets",
    "Insurance", "Training & Education", "Bank & Interest Charges",
    "Shipping & Delivery", "Subcontractors", "Other",
]


def main():
    parser = argparse.ArgumentParser(description="Add a manual expense entry")
    parser.add_argument("--vendor", required=True, help="Vendor/merchant name")
    parser.add_argument("--total", required=True, type=float, help="Total amount")
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"), help="Date (YYYY-MM-DD)")
    parser.add_argument("--category", default="Other", help="CRA expense category")
    parser.add_argument("--description", default="", help="Description of expense")
    parser.add_argument("--subtotal", type=float, default=None, help="Subtotal before tax")
    parser.add_argument("--gst", type=float, default=0, help="GST/HST amount")
    parser.add_argument("--pst", type=float, default=0, help="PST amount")
    parser.add_argument("--currency", default="CAD", help="Currency code")
    parser.add_argument("--payment", default="", help="Payment method")
    parser.add_argument("--source", default="Manual entry", help="Source of expense (e.g. 'Email receipt', 'Manual entry')")
    parser.add_argument("--confirmed", action="store_true", help="Actually save (without this, only shows preview)")
    args = parser.parse_args()

    # Auto-calculate subtotal if not provided
    subtotal = args.subtotal if args.subtotal is not None else args.total - args.gst - args.pst

    # Validate category
    if args.category not in CRA_CATEGORIES:
        # Try fuzzy match
        lower = args.category.lower()
        matched = [c for c in CRA_CATEGORIES if lower in c.lower()]
        if matched:
            args.category = matched[0]
        else:
            print(json.dumps({
                "error": f"Unknown category: {args.category}",
                "valid_categories": CRA_CATEGORIES
            }, indent=2))
            sys.exit(1)

    entry = {
        "date": args.date,
        "vendor": args.vendor,
        "description": args.description,
        "category": args.category,
        "subtotal": subtotal,
        "tax_gst_hst": args.gst,
        "tax_pst": args.pst,
        "total": args.total,
        "currency": args.currency,
        "payment_method": args.payment,
        "source": args.source,
    }

    # Check for duplicates
    dup_warning = None
    try:
        sys.path.insert(0, "/home/tonygale/openclaw/skills/receipts/scripts")
        from dedup_check import find_duplicate
        dup = find_duplicate(args.vendor, args.date, args.total)
        if dup:
            dup_warning = f"Possible duplicate: {dup['vendor']} on {dup['date']} for ${dup['total']:.2f} (row {dup['row']}, category: {dup['category']}). Already in spreadsheet."
    except Exception:
        pass

    if not args.confirmed:
        output = {"preview": entry, "next_step": "Review the entry above. Re-run with --confirmed to save it."}
        if dup_warning:
            output["duplicate_warning"] = dup_warning
        print(json.dumps(output, indent=2))
        sys.exit(0)

    # Save to Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed. Run: pip install --break-system-packages openpyxl"}))
        sys.exit(1)

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    os.makedirs(LOCAL_DIR, exist_ok=True)

    # Download existing Excel from Drive if not local
    if not os.path.exists(LOCAL_EXCEL):
        subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy",
                         DRIVE_ACCOUNTANT + EXCEL_NAME, LOCAL_DIR], capture_output=True)

    if os.path.exists(LOCAL_EXCEL):
        wb = openpyxl.load_workbook(LOCAL_EXCEL)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        headers = ["Date", "Vendor", "Description", "Category", "Subtotal", "GST/HST",
                    "PST", "Total", "Currency", "Payment Method", "Receipt File"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        ws2 = wb.create_sheet("Category Summary")
        headers2 = ["Category", "Total Spent", "GST/HST Paid", "PST Paid", "Transaction Count"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for row_num, cat in enumerate(CRA_CATEGORIES, 2):
            ws2.cell(row=row_num, column=1, value=cat)

    ws = wb["Transactions"]
    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=entry["date"]).border = THIN_BORDER
    ws.cell(row=row, column=2, value=entry["vendor"]).border = THIN_BORDER
    ws.cell(row=row, column=3, value=entry["description"]).border = THIN_BORDER
    ws.cell(row=row, column=4, value=entry["category"]).border = THIN_BORDER
    ws.cell(row=row, column=5, value=entry["subtotal"]).border = THIN_BORDER
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=entry["tax_gst_hst"]).border = THIN_BORDER
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=7, value=entry["tax_pst"]).border = THIN_BORDER
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    ws.cell(row=row, column=8, value=entry["total"]).border = THIN_BORDER
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    ws.cell(row=row, column=9, value=entry["currency"]).border = THIN_BORDER
    ws.cell(row=row, column=10, value=entry["payment_method"]).border = THIN_BORDER
    ws.cell(row=row, column=11, value=entry["source"]).border = THIN_BORDER

    wb.save(LOCAL_EXCEL)

    # Upload to Drive
    subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy", LOCAL_EXCEL, DRIVE_ACCOUNTANT],
                    capture_output=True)

    print(json.dumps({
        "saved": True,
        "entry": entry,
        "message": f"Expense saved to {EXCEL_NAME} and synced to Google Drive."
    }, indent=2))


if __name__ == "__main__":
    main()
