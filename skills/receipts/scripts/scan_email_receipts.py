#!/usr/bin/env python3
"""Scan Gmail for receipt emails and extract expense data.

Searches for common receipt/invoice emails and uses Gemini to extract
expense details, then saves them to the shared Expenses spreadsheet.

Tracks processed emails in a local JSON file to avoid re-processing
and wasting Gemini API tokens.

Usage:
    python3 scan_email_receipts.py                    # Scan last 7 days (new only)
    python3 scan_email_receipts.py --days 30          # Scan last 30 days (new only)
    python3 scan_email_receipts.py --from uber        # Only from Uber
    python3 scan_email_receipts.py --list             # List found receipts without saving
    python3 scan_email_receipts.py --save-all         # Save all found receipts
    python3 scan_email_receipts.py --reprocess        # Ignore tracker, reprocess all
"""

import argparse
import email
import hashlib
import imaplib
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timedelta
from email.header import decode_header

RCLONE_CONFIG = "/data/.config/rclone/rclone.conf"
LOCAL_DIR = "/data/.openclaw/workspace/receipts"
YEAR = datetime.now().strftime("%Y")
EXCEL_NAME = f"Expenses_{YEAR}.xlsx"
LOCAL_EXCEL = os.path.join(LOCAL_DIR, EXCEL_NAME)
DRIVE_ACCOUNTANT = "gdrive:Accountant/"

# Tracker file — stores which emails have already been processed
TRACKER_FILE = "/data/.openclaw/workspace/receipts/processed_emails.json"

GMAIL_USER = os.environ.get("GMAIL_USER", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

# Common receipt sender patterns
RECEIPT_SENDERS = [
    "uber", "lyft", "amazon", "doordash", "skip", "grubhub",
    "netflix", "spotify", "apple", "google", "microsoft",
    "shopify", "paypal", "stripe", "square", "interac",
    "costco", "walmart", "bestbuy", "staples", "home depot",
    "tim hortons", "starbucks", "mcdonalds",
]

RECEIPT_SUBJECT_KEYWORDS = [
    "receipt", "invoice", "order confirmation", "payment confirmation",
    "your order", "transaction", "billing statement", "purchase",
    "e-receipt", "digital receipt",
]

# Tony's email addresses — forwarded emails from these are treated as potential receipts
TONY_FORWARDING_ADDRESSES = [
    "tonygale11@gmail.com",
    "tonygale24@gmail.com",
    "tonygale@myprojectworld.ca",
    "tonygale@galehomedesigns.com",
    "decadesdevelopments@gmail.com",
]

CRA_CATEGORIES = [
    "Advertising & Marketing", "Meals & Entertainment", "Office Supplies",
    "Professional Fees", "Rent & Lease", "Telephone & Internet", "Travel",
    "Vehicle Expenses", "Software & Subscriptions", "Equipment & Assets",
    "Insurance", "Training & Education", "Bank & Interest Charges",
    "Shipping & Delivery", "Subcontractors", "Other",
]


# --- Processed email tracker ---

def load_tracker():
    """Load the set of already-processed email fingerprints."""
    if os.path.exists(TRACKER_FILE):
        try:
            with open(TRACKER_FILE) as f:
                data = json.load(f)
                return data
        except Exception:
            pass
    return {"processed": {}, "stats": {"total_scanned": 0, "total_receipts": 0, "total_skipped": 0}}


def save_tracker(tracker):
    """Save tracker to disk."""
    os.makedirs(os.path.dirname(TRACKER_FILE), exist_ok=True)
    with open(TRACKER_FILE, "w") as f:
        json.dump(tracker, f, indent=2)


def email_fingerprint(message_id, from_addr, subject, date_str):
    """Create a unique fingerprint for an email to detect duplicates."""
    # Use Message-ID if available (most reliable), otherwise hash from+subject+date
    if message_id:
        return hashlib.sha256(message_id.encode()).hexdigest()[:16]
    raw = f"{from_addr}|{subject}|{date_str}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


# --- Email parsing ---

def decode_mime_header(header_value):
    """Decode MIME encoded header."""
    if not header_value:
        return ""
    decoded_parts = decode_header(header_value)
    result = []
    for part, charset in decoded_parts:
        if isinstance(part, bytes):
            result.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            result.append(part)
    return " ".join(result)


def strip_html(html_text):
    """Strip HTML tags and decode entities."""
    text = re.sub(r'<style[^>]*>.*?</style>', '', html_text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&lt;', '<', text)
    text = re.sub(r'&gt;', '>', text)
    text = re.sub(r'&#\d+;', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def get_email_body(msg):
    """Extract text body from email message."""
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            if content_type == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or "utf-8"
                    body = payload.decode(charset, errors="replace")
                    break
            elif content_type == "text/html" and not body:
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or "utf-8"
                    body = strip_html(payload.decode(charset, errors="replace"))
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or "utf-8"
            if msg.get_content_type() == "text/html":
                body = strip_html(payload.decode(charset, errors="replace"))
            else:
                body = payload.decode(charset, errors="replace")
    return body[:4000]  # Limit to 4000 chars for Gemini


# --- Receipt detection ---

def is_forwarded_from_tony(from_addr):
    """Check if email is from one of Tony's forwarding addresses."""
    from_lower = from_addr.lower()
    return any(addr in from_lower for addr in TONY_FORWARDING_ADDRESSES)


def is_forwarded_email(subject):
    """Check if subject indicates a forwarded email."""
    stripped = subject.strip().lower()
    return stripped.startswith("fwd:") or stripped.startswith("fw:")


def is_receipt_email(from_addr, subject):
    """Check if an email is likely a receipt."""
    from_lower = from_addr.lower()
    subject_lower = subject.lower()

    # Forwarded emails from Tony's addresses are always potential receipts
    # (Tony only forwards receipts/invoices to the business account)
    if is_forwarded_from_tony(from_addr) and is_forwarded_email(subject):
        return True

    # Check sender against known receipt senders
    for sender in RECEIPT_SENDERS:
        if sender in from_lower:
            return True

    # Check subject (including after stripping Fwd: prefix)
    clean_subject = re.sub(r'^(fwd?|fw):\s*', '', subject_lower, flags=re.IGNORECASE).strip()
    for keyword in RECEIPT_SUBJECT_KEYWORDS:
        if keyword in subject_lower or keyword in clean_subject:
            return True

    return False


# --- Gemini extraction ---

def extract_expense_with_gemini(email_text, from_addr, subject, date_str):
    """Use Gemini to extract expense data from email text."""
    if not GEMINI_API_KEY:
        return None

    prompt = f"""Extract expense/receipt information from this email. Return ONLY valid JSON with these fields:
{{
    "is_receipt": true/false,
    "vendor": "company name",
    "description": "what was purchased",
    "date": "{date_str}",
    "subtotal": 0.00,
    "tax_gst_hst": 0.00,
    "tax_pst": 0.00,
    "total": 0.00,
    "currency": "CAD",
    "payment_method": "card type or method",
    "category": "one of: {', '.join(CRA_CATEGORIES)}"
}}

If this is NOT a receipt/invoice (e.g. it's a marketing email, newsletter, etc.), set is_receipt to false.
Use the email date ({date_str}) if no specific transaction date is found.
For Canadian taxes: GST/HST is federal (5-15%), PST is provincial (0-10%).
If tax amounts aren't specified, set them to 0.

Email from: {from_addr}
Subject: {subject}
Body:
{email_text}"""

    import urllib.request
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}"

    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
    }).encode()

    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
            text = result["candidates"][0]["content"]["parts"][0]["text"]
            # Strip markdown code blocks
            text = re.sub(r'^```json\s*', '', text.strip())
            text = re.sub(r'\s*```$', '', text.strip())
            return json.loads(text)
    except Exception as e:
        print(f"  Gemini error: {e}", file=sys.stderr)
        return None


# --- Gmail scanning ---

def scan_gmail(days=7, sender_filter=None):
    """Scan Gmail for receipt emails."""
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        print(json.dumps({"error": "GMAIL_USER and GMAIL_APP_PASSWORD not set"}))
        sys.exit(1)

    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
    mail.select("INBOX")

    since_date = (datetime.now() - timedelta(days=days)).strftime("%d-%b-%Y")
    search_criteria = f'(SINCE {since_date})'

    if sender_filter:
        search_criteria = f'(SINCE {since_date} FROM "{sender_filter}")'

    _, msg_ids = mail.search(None, search_criteria)
    if not msg_ids[0]:
        mail.logout()
        return []

    receipts = []
    ids = msg_ids[0].split()

    for msg_id in ids:
        _, msg_data = mail.fetch(msg_id, "(RFC822)")
        raw = msg_data[0][1]
        msg = email.message_from_bytes(raw)

        from_addr = decode_mime_header(msg.get("From", ""))
        subject = decode_mime_header(msg.get("Subject", ""))
        date_str = msg.get("Date", "")
        message_id = msg.get("Message-ID", "")

        # Parse date
        try:
            from email.utils import parsedate_to_datetime
            email_date = parsedate_to_datetime(date_str)
            date_formatted = email_date.strftime("%Y-%m-%d")
        except Exception:
            date_formatted = datetime.now().strftime("%Y-%m-%d")

        if not is_receipt_email(from_addr, subject):
            continue

        fp = email_fingerprint(message_id, from_addr, subject, date_str)
        body = get_email_body(msg)
        receipts.append({
            "fingerprint": fp,
            "message_id": message_id,
            "from": from_addr,
            "subject": subject,
            "date": date_formatted,
            "body": body,
        })

    mail.logout()
    return receipts


# --- Excel saving ---

def save_expense(entry):
    """Save an expense entry to the Excel spreadsheet."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return {"error": "openpyxl not installed"}

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    os.makedirs(LOCAL_DIR, exist_ok=True)

    if not os.path.exists(LOCAL_EXCEL):
        subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy",
                         DRIVE_ACCOUNTANT + EXCEL_NAME, LOCAL_DIR], capture_output=True)

    if os.path.exists(LOCAL_EXCEL):
        wb = openpyxl.load_workbook(LOCAL_EXCEL)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        headers = ["Date", "Vendor", "Description", "Category", "Subtotal", "GST/HST",
                    "PST", "Total", "Currency", "Payment Method", "Receipt File"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        ws2 = wb.create_sheet("Category Summary")
        headers2 = ["Category", "Total Spent", "GST/HST Paid", "PST Paid", "Transaction Count"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for row_num, cat in enumerate(CRA_CATEGORIES, 2):
            ws2.cell(row=row_num, column=1, value=cat)

    ws = wb["Transactions"]
    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=entry.get("date", "")).border = THIN_BORDER
    ws.cell(row=row, column=2, value=entry.get("vendor", "")).border = THIN_BORDER
    ws.cell(row=row, column=3, value=entry.get("description", "")).border = THIN_BORDER
    ws.cell(row=row, column=4, value=entry.get("category", "Other")).border = THIN_BORDER
    ws.cell(row=row, column=5, value=entry.get("subtotal", 0)).border = THIN_BORDER
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=entry.get("tax_gst_hst", 0)).border = THIN_BORDER
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=7, value=entry.get("tax_pst", 0)).border = THIN_BORDER
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    ws.cell(row=row, column=8, value=entry.get("total", 0)).border = THIN_BORDER
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    ws.cell(row=row, column=9, value=entry.get("currency", "CAD")).border = THIN_BORDER
    ws.cell(row=row, column=10, value=entry.get("payment_method", "")).border = THIN_BORDER
    ws.cell(row=row, column=11, value="Email receipt").border = THIN_BORDER

    wb.save(LOCAL_EXCEL)
    return {"saved": True}


# --- Main ---

def main():
    parser = argparse.ArgumentParser(description="Scan Gmail for receipt emails")
    parser.add_argument("--days", type=int, default=7, help="Number of days to scan back (default: 7)")
    parser.add_argument("--from", dest="sender", help="Filter by sender name/address")
    parser.add_argument("--list", action="store_true", help="List found receipts without saving")
    parser.add_argument("--save-all", action="store_true", help="Save all found receipts")
    parser.add_argument("--reprocess", action="store_true", help="Ignore tracker, reprocess all emails")
    args = parser.parse_args()

    tracker = load_tracker()

    print(f"Scanning Gmail for receipts (last {args.days} days)...", file=sys.stderr)
    receipts = scan_gmail(days=args.days, sender_filter=args.sender)

    if not receipts:
        print(json.dumps({"found": 0, "new": 0, "message": "No receipt emails found."}))
        return

    # Filter out already-processed emails (unless --reprocess)
    new_receipts = []
    already_processed = 0
    for r in receipts:
        fp = r["fingerprint"]
        if not args.reprocess and fp in tracker["processed"]:
            already_processed += 1
            continue
        new_receipts.append(r)

    if not new_receipts:
        print(json.dumps({
            "found": len(receipts),
            "new": 0,
            "already_processed": already_processed,
            "message": "All receipt emails have already been processed. Use --reprocess to force re-scan."
        }))
        return

    print(f"  {len(new_receipts)} new emails to process ({already_processed} already processed)", file=sys.stderr)

    results = []
    for r in new_receipts:
        print(f"  Processing: {r['subject'][:60]}...", file=sys.stderr)
        expense_data = extract_expense_with_gemini(r["body"], r["from"], r["subject"], r["date"])

        fp = r["fingerprint"]

        if not expense_data or not expense_data.get("is_receipt"):
            results.append({
                "from": r["from"],
                "subject": r["subject"],
                "date": r["date"],
                "status": "skipped",
                "reason": "Not a receipt or couldn't extract data"
            })
            # Still mark as processed so we don't waste tokens re-checking it
            tracker["processed"][fp] = {
                "date_scanned": datetime.now().isoformat(),
                "subject": r["subject"][:80],
                "result": "not_a_receipt"
            }
            tracker["stats"]["total_skipped"] = tracker["stats"].get("total_skipped", 0) + 1
            continue

        entry = {
            "from": r["from"],
            "subject": r["subject"],
            "date": expense_data.get("date", r["date"]),
            "vendor": expense_data.get("vendor", "Unknown"),
            "description": expense_data.get("description", ""),
            "category": expense_data.get("category", "Other"),
            "subtotal": expense_data.get("subtotal", 0),
            "tax_gst_hst": expense_data.get("tax_gst_hst", 0),
            "tax_pst": expense_data.get("tax_pst", 0),
            "total": expense_data.get("total", 0),
            "currency": expense_data.get("currency", "CAD"),
            "payment_method": expense_data.get("payment_method", ""),
            "status": "found",
        }

        # Check for duplicates in spreadsheet
        try:
            sys.path.insert(0, "/data/skills/receipts/scripts")
            from dedup_check import find_duplicate
            dup = find_duplicate(entry.get("vendor"), entry.get("date"), entry.get("total"))
            if dup:
                entry["status"] = "duplicate"
                entry["duplicate_of"] = f"{dup['vendor']} on {dup['date']} for ${dup['total']:.2f} (row {dup['row']})"
                tracker["processed"][fp] = {
                    "date_scanned": datetime.now().isoformat(),
                    "subject": r["subject"][:80],
                    "result": "duplicate",
                    "vendor": entry["vendor"],
                    "total": entry["total"],
                }
                results.append(entry)
                continue
        except Exception:
            pass

        if args.save_all:
            save_result = save_expense(entry)
            entry["status"] = "saved" if save_result.get("saved") else "error"
            if entry["status"] == "saved":
                tracker["processed"][fp] = {
                    "date_scanned": datetime.now().isoformat(),
                    "subject": r["subject"][:80],
                    "result": "saved",
                    "vendor": entry["vendor"],
                    "total": entry["total"],
                }
                tracker["stats"]["total_receipts"] = tracker["stats"].get("total_receipts", 0) + 1
        elif args.list:
            entry["status"] = "preview"
            # Don't mark as processed on --list (preview only)

        results.append(entry)

    tracker["stats"]["total_scanned"] = tracker["stats"].get("total_scanned", 0) + len(new_receipts)

    # Save tracker (always, even for skipped emails — prevents re-processing)
    save_tracker(tracker)

    # Upload to Drive if we saved anything
    if args.save_all and any(r.get("status") == "saved" for r in results):
        subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy", LOCAL_EXCEL, DRIVE_ACCOUNTANT],
                        capture_output=True)

    output = {
        "found": len(receipts),
        "new": len(new_receipts),
        "already_processed": already_processed,
        "saved": sum(1 for r in results if r.get("status") == "saved"),
        "skipped": sum(1 for r in results if r.get("status") == "skipped"),
        "receipts": results,
    }

    if not args.save_all and not args.list:
        output["next_step"] = "Use --list to preview or --save-all to save all found receipts."

    print(json.dumps(output, indent=2))


if __name__ == "__main__":
    main()
