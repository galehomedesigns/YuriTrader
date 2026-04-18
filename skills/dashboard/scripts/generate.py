#!/usr/bin/env python3
"""Generate the Tony Gale / Decades Developments dashboard HTML."""

import json
import os
import subprocess
import sys
from datetime import datetime

RCLONE_CONFIG = "/data/.config/rclone/rclone.conf"
LOCAL_DIR = "/data/.openclaw/workspace/receipts"
YEAR = datetime.now().strftime("%Y")
EXCEL_NAME = f"Expenses_{YEAR}.xlsx"
LOCAL_EXCEL = os.path.join(LOCAL_DIR, EXCEL_NAME)
DRIVE_ACCOUNTANT = "gdrive:Accountant/"
DATA_FILE = "/data/.openclaw/workspace/dashboard/data.json"
OUTPUT_FILE = "/data/.openclaw/canvas/dashboard.html"


def load_expense_data():
    """Load expense data from Excel file."""
    # Download latest from Drive
    os.makedirs(LOCAL_DIR, exist_ok=True)
    subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy",
                     DRIVE_ACCOUNTANT + EXCEL_NAME, LOCAL_DIR], capture_output=True)

    if not os.path.exists(LOCAL_EXCEL):
        return {"transactions": [], "by_category": {}, "totals": {"spent": 0, "gst": 0, "pst": 0, "count": 0}}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(LOCAL_EXCEL)
        ws = wb["Transactions"]
    except Exception:
        return {"transactions": [], "by_category": {}, "totals": {"spent": 0, "gst": 0, "pst": 0, "count": 0}}

    transactions = []
    by_category = {}
    total_spent = 0
    total_gst = 0
    total_pst = 0

    for row in range(2, ws.max_row + 1):
        date = ws.cell(row=row, column=1).value or ""
        vendor = ws.cell(row=row, column=2).value or ""
        desc = ws.cell(row=row, column=3).value or ""
        category = ws.cell(row=row, column=4).value or "Other"
        subtotal = ws.cell(row=row, column=5).value or 0
        gst = ws.cell(row=row, column=6).value or 0
        pst = ws.cell(row=row, column=7).value or 0
        total = ws.cell(row=row, column=8).value or 0

        if not vendor and not total:
            continue

        transactions.append({
            "date": str(date), "vendor": vendor, "description": desc,
            "category": category, "total": total, "gst": gst, "pst": pst
        })

        if category not in by_category:
            by_category[category] = {"total": 0, "count": 0}
        by_category[category]["total"] += total
        by_category[category]["count"] += 1
        total_spent += total
        total_gst += gst
        total_pst += pst

    return {
        "transactions": transactions[-20:],  # Last 20
        "by_category": by_category,
        "totals": {"spent": total_spent, "gst": total_gst, "pst": total_pst, "count": len(transactions)}
    }


def load_dashboard_data():
    """Load todos and priorities."""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE) as f:
            return json.load(f)
    return {"todos": [], "priorities": [], "notes": ""}


def generate_category_bars(by_category, total_spent):
    """Generate HTML for category spending bars."""
    if not by_category:
        return '<p style="color:#8899aa;">No expenses recorded yet.</p>'

    sorted_cats = sorted(by_category.items(), key=lambda x: x[1]["total"], reverse=True)
    max_val = sorted_cats[0][1]["total"] if sorted_cats else 1
    html = ""
    colors = ["#3498db", "#2ecc71", "#e74c3c", "#f39c12", "#9b59b6", "#1abc9c",
              "#e67e22", "#34495e", "#16a085", "#c0392b", "#2980b9", "#8e44ad"]

    for i, (cat, data) in enumerate(sorted_cats):
        pct = (data["total"] / max_val * 100) if max_val > 0 else 0
        color = colors[i % len(colors)]
        html += f'''
        <div style="margin-bottom:12px;">
            <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
                <span style="color:#ccd6dd;font-size:14px;">{cat}</span>
                <span style="color:#8899aa;font-size:13px;">${data["total"]:,.2f} ({data["count"]})</span>
            </div>
            <div style="background:#1a2332;border-radius:6px;overflow:hidden;height:8px;">
                <div style="background:{color};width:{pct:.1f}%;height:100%;border-radius:6px;transition:width 0.5s;"></div>
            </div>
        </div>'''
    return html


def generate_todos_html(todos):
    """Generate HTML for todo list."""
    if not todos:
        return '<p style="color:#8899aa;">No tasks yet. Tell Yuri to add some!</p>'

    html = ""
    priority_colors = {"high": "#e74c3c", "medium": "#f39c12", "low": "#3498db"}
    status_icons = {"pending": "○", "in_progress": "◐", "done": "●"}

    for todo in todos:
        color = priority_colors.get(todo.get("priority", "medium"), "#8899aa")
        icon = status_icons.get(todo.get("status", "pending"), "○")
        done_style = "text-decoration:line-through;opacity:0.5;" if todo.get("status") == "done" else ""
        due = f' <span style="color:#8899aa;font-size:12px;">Due: {todo["due"]}</span>' if todo.get("due") else ""
        html += f'''
        <div style="display:flex;align-items:center;padding:10px 0;border-bottom:1px solid #1a2332;">
            <span style="color:{color};font-size:18px;margin-right:12px;">{icon}</span>
            <div style="{done_style}">
                <span style="color:#ccd6dd;font-size:14px;">{todo["text"]}</span>{due}
            </div>
        </div>'''
    return html


def generate_priorities_html(priorities):
    """Generate HTML for priorities."""
    if not priorities:
        return '<p style="color:#8899aa;">No priorities set.</p>'

    html = ""
    icons = {"Marketing": "📣", "Platform": "⚙️", "Operations": "📋", "Networking": "🤝"}
    for p in priorities:
        icon = icons.get(p["category"], "📌")
        items_html = "".join(f'<li style="color:#8899aa;font-size:13px;margin:4px 0;">{item}</li>' for item in p["items"])
        html += f'''
        <div style="margin-bottom:16px;">
            <div style="color:#ccd6dd;font-size:15px;font-weight:600;margin-bottom:6px;">{icon} {p["category"]}</div>
            <ul style="margin:0;padding-left:20px;">{items_html}</ul>
        </div>'''
    return html


def generate_recent_transactions(transactions):
    """Generate HTML table for recent transactions."""
    if not transactions:
        return '<p style="color:#8899aa;">No transactions yet.</p>'

    rows = ""
    for t in reversed(transactions[-10:]):
        rows += f'''
        <tr>
            <td style="padding:8px 12px;color:#8899aa;font-size:13px;">{t["date"]}</td>
            <td style="padding:8px 12px;color:#ccd6dd;font-size:13px;">{t["vendor"][:30]}</td>
            <td style="padding:8px 12px;color:#ccd6dd;font-size:13px;">${t["total"]:,.2f}</td>
            <td style="padding:8px 12px;color:#8899aa;font-size:13px;">{t["category"]}</td>
        </tr>'''

    return f'''
    <table style="width:100%;border-collapse:collapse;">
        <thead>
            <tr style="border-bottom:2px solid #1a2332;">
                <th style="padding:8px 12px;text-align:left;color:#667788;font-size:12px;text-transform:uppercase;">Date</th>
                <th style="padding:8px 12px;text-align:left;color:#667788;font-size:12px;text-transform:uppercase;">Vendor</th>
                <th style="padding:8px 12px;text-align:left;color:#667788;font-size:12px;text-transform:uppercase;">Amount</th>
                <th style="padding:8px 12px;text-align:left;color:#667788;font-size:12px;text-transform:uppercase;">Category</th>
            </tr>
        </thead>
        <tbody>{rows}</tbody>
    </table>'''


def main():
    now = datetime.now()
    expenses = load_expense_data()
    data = load_dashboard_data()

    totals = expenses["totals"]
    category_bars = generate_category_bars(expenses["by_category"], totals["spent"])
    todos_html = generate_todos_html(data.get("todos", []))
    priorities_html = generate_priorities_html(data.get("priorities", []))
    transactions_html = generate_recent_transactions(expenses["transactions"])

    month_name = now.strftime("%B %Y")

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Decades Developments — Dashboard</title>
    <style>
        * {{ margin:0; padding:0; box-sizing:border-box; }}
        body {{ background:#0d1117; color:#ccd6dd; font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; padding:20px; }}
        .header {{ text-align:center; padding:30px 0 20px; border-bottom:1px solid #1a2332; margin-bottom:30px; }}
        .header h1 {{ font-size:28px; color:#ffffff; font-weight:700; }}
        .header .subtitle {{ color:#8899aa; font-size:14px; margin-top:6px; }}
        .header .updated {{ color:#556677; font-size:12px; margin-top:4px; }}
        .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:24px; max-width:1200px; margin:0 auto; }}
        .card {{ background:#161b22; border:1px solid #21262d; border-radius:12px; padding:24px; }}
        .card h2 {{ font-size:16px; color:#ffffff; margin-bottom:16px; padding-bottom:10px; border-bottom:1px solid #1a2332; }}
        .stat-row {{ display:flex; gap:16px; margin-bottom:24px; max-width:1200px; margin:0 auto 24px; }}
        .stat {{ flex:1; background:#161b22; border:1px solid #21262d; border-radius:12px; padding:20px; text-align:center; }}
        .stat .value {{ font-size:32px; font-weight:700; color:#ffffff; }}
        .stat .label {{ font-size:13px; color:#8899aa; margin-top:4px; }}
        .full-width {{ grid-column:1/-1; }}
        .refresh-btn {{
            display:inline-flex; align-items:center; gap:6px;
            background:#161b22; border:1px solid #21262d; border-radius:8px;
            color:#8899aa; padding:8px 16px; font-size:13px; cursor:pointer;
            transition:all 0.2s; margin-top:10px;
        }}
        .refresh-btn:hover {{ background:#1f2937; color:#ccd6dd; border-color:#3498db; }}
        .refresh-btn.spinning svg {{ animation:spin 1s linear infinite; }}
        @keyframes spin {{ from {{ transform:rotate(0deg); }} to {{ transform:rotate(360deg); }} }}
        @media (max-width:768px) {{
            .grid {{ grid-template-columns:1fr; }}
            .stat-row {{ flex-direction:column; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Decades Developments</h1>
        <div class="subtitle">Tony Gale, P.Tech, PMP — The Project Wheel</div>
        <div class="updated">Last updated: {now.strftime("%B %d, %Y at %I:%M %p ET")}</div>
        <button class="refresh-btn" onclick="refreshDashboard(this)">
            <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21.5 2v6h-6M2.5 22v-6h6M2 11.5a10 10 0 0 1 18.8-4.3M22 12.5a10 10 0 0 1-18.8 4.2"/></svg>
            Refresh
        </button>
    </div>
    <script>
    function refreshDashboard(btn) {{
        btn.classList.add('spinning');
        btn.querySelector('svg').style.color = '#3498db';
        setTimeout(() => location.reload(), 300);
    }}
    </script>

    <div class="stat-row">
        <div class="stat">
            <div class="value" style="color:#e74c3c;">${totals["spent"]:,.2f}</div>
            <div class="label">Total Spent ({month_name})</div>
        </div>
        <div class="stat">
            <div class="value" style="color:#f39c12;">${totals["gst"]:,.2f}</div>
            <div class="label">GST/HST Paid</div>
        </div>
        <div class="stat">
            <div class="value" style="color:#3498db;">${totals["pst"]:,.2f}</div>
            <div class="label">PST Paid</div>
        </div>
        <div class="stat">
            <div class="value" style="color:#2ecc71;">{totals["count"]}</div>
            <div class="label">Receipts Processed</div>
        </div>
    </div>

    <div class="grid">
        <div class="card">
            <h2>Spending by Category</h2>
            {category_bars}
        </div>

        <div class="card">
            <h2>To-Do List</h2>
            {todos_html}
        </div>

        <div class="card">
            <h2>Priorities</h2>
            {priorities_html}
        </div>

        <div class="card">
            <h2>Recent Transactions</h2>
            {transactions_html}
        </div>
    </div>

    <div style="text-align:center;padding:30px 0;color:#556677;font-size:12px;">
        Powered by Yuri — Decades Developments AI Assistant
    </div>
</body>
</html>'''

    with open(OUTPUT_FILE, "w") as f:
        f.write(html)

    # Also save a standalone copy accessible without OpenClaw login
    standalone_copy = "/data/dashboard.html"
    with open(standalone_copy, "w") as f:
        f.write(html)

    # Sync to Google Drive so Tony can access from any device
    subprocess.run(["rclone", "--config", RCLONE_CONFIG, "copy",
                     standalone_copy, "gdrive:Dashboard/"], capture_output=True)

    print(f"Dashboard generated: {OUTPUT_FILE}")
    print(f"Expenses: ${totals['spent']:,.2f} across {totals['count']} receipts")
    print(f"Todos: {len(data.get('todos', []))} items")
    print(f"Synced to Google Drive: Dashboard/dashboard.html")
    print(f"Updated: {now.strftime('%Y-%m-%d %I:%M %p')}")


if __name__ == "__main__":
    main()
