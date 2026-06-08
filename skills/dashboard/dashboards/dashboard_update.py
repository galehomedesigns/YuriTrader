#!/usr/bin/env python3
"""Build dashboard.html (personal — expenses, todos, priorities).

Reads ~/openclaw/state/dashboard/data.json as the source of truth.
Optionally pulls the latest Expenses_<year>.xlsx via rclone from the
`gdrive:Accountant/` remote and folds its by-category aggregates in,
but degrades gracefully if rclone isn't configured.
"""
from __future__ import annotations

import json
import subprocess
from datetime import datetime, timezone
from pathlib import Path

SKILL_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = Path(__file__).resolve().parent / "dashboard.template.html"
OUT = Path("/home/tonygale/openclaw/canvas/dashboard.html")
STATE_DIR = Path("/home/tonygale/openclaw/state/dashboard")
STATE_FILE = STATE_DIR / "data.json"


def load_state() -> dict:
    if not STATE_FILE.exists():
        return {}
    try:
        return json.loads(STATE_FILE.read_text())
    except Exception:
        return {}


def _date_key(v) -> str:
    """Normalize an Excel date cell (datetime or str) to an ISO 'YYYY-MM-DD'
    string for both display and lexical sorting."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def try_refresh_excel() -> tuple[dict, str | None]:
    """Best-effort pull of the current Excel ledger via rclone — this is the
    SOURCE OF TRUTH for expenses (it's the workbook the accountant receives).

    Returns ({by_category, total, count, recent}, warning|None). On any failure
    returns ({}, warning) so build_data() can fall back to cached data.json.
    """
    year = datetime.now().strftime("%Y")
    excel_name = f"Expenses_{year}.xlsx"
    local = STATE_DIR / excel_name

    try:
        STATE_DIR.mkdir(parents=True, exist_ok=True)
        result = subprocess.run(
            ["rclone", "copy", f"gdrive:Accountant/{excel_name}", str(STATE_DIR)],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode != 0:
            return {}, f"rclone fetch failed (rc={result.returncode}); showing cached expense data only."
    except FileNotFoundError:
        return {}, "rclone not installed on GX10 — showing cached expense data only."
    except subprocess.TimeoutExpired:
        return {}, "rclone timed out — showing cached expense data only."

    if not local.exists():
        return {}, f"Expense workbook {excel_name} not found on gdrive — showing cached data only."

    try:
        import openpyxl  # type: ignore
    except ImportError:
        return {}, "openpyxl not installed in venv — showing cached expense data only."

    try:
        wb = openpyxl.load_workbook(str(local), data_only=True)
        ws = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active
    except Exception as e:
        return {}, f"Could not read {excel_name}: {type(e).__name__}."

    try:
        header = [c.value for c in ws[1]]
        def idx(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None
        date_idx = idx("Date")
        vendor_idx = idx("Vendor")
        desc_idx = idx("Description")
        cat_idx = idx("Category")
        amt_idx = idx("Amount", "Total")
        if cat_idx is None or amt_idx is None or date_idx is None:
            return {}, "Expense workbook has unexpected columns — showing cached data only."

        by_category: dict[str, float] = {}
        total = 0.0
        txns = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip blank/trailing rows: a real transaction has a date.
            if date_idx >= len(row) or row[date_idx] in (None, ""):
                continue
            try:
                amt = float(row[amt_idx] or 0)
            except (ValueError, TypeError):
                continue
            cat = row[cat_idx] or "Uncategorized"
            by_category[cat] = round(by_category.get(cat, 0.0) + amt, 2)
            total += amt
            txns.append({
                "date": _date_key(row[date_idx]),
                "vendor": row[vendor_idx] if vendor_idx is not None else "",
                "amount": round(amt, 2),
                "description": row[desc_idx] if desc_idx is not None else "",
            })
    except Exception as e:
        return {}, f"Could not parse {excel_name}: {type(e).__name__}."

    recent = sorted(txns, key=lambda t: t["date"], reverse=True)[:8]
    return {
        "by_category": by_category,
        "total": round(total, 2),
        "count": len(txns),
        "recent": recent,
    }, None


def build_data() -> dict:
    state = load_state()
    excel, warning = try_refresh_excel()

    # The Excel ledger is authoritative; fall back to cached data.json only
    # when it can't be read (rclone/openpyxl/file missing).
    by_category = excel.get("by_category") or state.get("by_category") or {}
    if excel:
        total_expenses = excel["total"]
        receipt_count = excel["count"]
        recent_expenses = excel["recent"]
    else:
        total_expenses = state.get("totalExpenses") or sum(by_category.values())
        receipt_count = state.get("receiptCount")
        recent_expenses = state.get("recentExpenses") or []

    todos = state.get("todos") or []
    active_todos = sum(1 for t in todos if t.get("status") != "done")
    high_priority = sum(1 for t in todos if t.get("priority") == "high" and t.get("status") != "done")

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "warning": warning,
        "summary": {
            "total_expenses": total_expenses,
            "receipt_count": receipt_count,
            "active_todos": active_todos,
            "high_priority": high_priority,
        },
        "by_category": by_category,
        "recent_expenses": recent_expenses,
        "todos": todos,
    }


def main() -> None:
    template = TEMPLATE.read_text()
    data = build_data()
    html = template.replace("{{DATA}}", json.dumps(data, default=str))
    OUT.write_text(html)
    print(
        f"Wrote {OUT} ({len(html)} bytes, "
        f"{len(data['by_category'])} categories, "
        f"{len(data['recent_expenses'])} recent txs, "
        f"{len(data['todos'])} todos)"
    )


if __name__ == "__main__":
    main()
